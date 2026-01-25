import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO

def exportar_excel_depreciacion(queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Depreciacion_Activos_{datetime.date.today()}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciacion Activos"

    # Encabezados
    headers = [
        'Código', 'Cuenta', 'Nombre Cuenta', 'Descripción', 
        'Fecha Adq.', 'Costo Inicial', 'Estado', 'Valor Libros'
    ]
    # Años dinámicos
    years = [f"Dep. {y}" for y in range(2005, 2027)]
    headers.extend(years)

    ws.append(headers)

    # Estilo encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    import openpyxl 
    
    # Datos
    for obj in queryset:
        row = [
            obj.codigo,
            obj.codcue,
            obj.nombre_cuenta,
            obj.descripcion,
            obj.fecha_adquisicion,
            obj.costo_inicial,
            obj.estado,
            obj.valor_en_libros_actual
        ]
        
        # Valores de depreciación
        dep_values = [
            getattr(obj, f'dep_{y}', 0) or 0 for y in range(2005, 2027)
        ]
        row.extend(dep_values)
        ws.append(row)

    wb.save(response)
    return response

def exportar_pdf_depreciacion(queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Depreciacion_Activos_{datetime.date.today()}.pdf"'

    buffer = BytesIO()
    # Usamos A3 Landscape para maximizar el ancho
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center

    elements.append(Paragraph(f"Reporte de Depreciación de Activos Fijos - {datetime.date.today()}", title_style))
    elements.append(Spacer(1, 20))

    # Encabezados reducidos para que quepa todo
    headers = ['Cod', 'Cuenta', 'Desc.', 'Fecha', 'Costo', 'V. Libros']
    # Años abreviados (05, 06...)
    years = [str(y)[2:] for y in range(2005, 2027)]
    headers.extend(years)

    data = [headers]

    for obj in queryset:
        row = [
            str(obj.codigo)[:10], # Truncar
            str(obj.codcue),
            str(obj.descripcion)[:20], # Truncar descripción
            str(obj.fecha_adquisicion),
            f"{obj.costo_inicial:.2f}",
            f"{obj.valor_en_libros_actual:.2f}"
        ]
        
        dep_values = [
            f"{getattr(obj, f'dep_{y}', 0) or 0:.2f}" for y in range(2005, 2027)
        ]
        row.extend(dep_values)
        data.append(row)

    # Tabla
    # Calcular anchos de columna dinámicos o fijos muy pequeños
    col_widths = [50, 40, 100, 60, 60, 60] + [35] * len(years) 
    
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005b96')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8), # Fuente pequeña encabezado
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7), # Fuente muy pequeña datos
        ('ALIGN', (2, 1), (2, -1), 'LEFT'), # Descripción a la izquierda
    ]))

    elements.append(t)
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response
